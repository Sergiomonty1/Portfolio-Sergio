from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from datetime import date

wb = Workbook()
ws = wb.active
ws.title = "Registro Apuestas"

headers = ["Fecha", "Evento", "Apuesta", "Cuota", "MontoInvertido", "Resultado", "Retorno", "GananciaNeta", "ROI"]
ws.append(headers)

sample = [
    [date(2026,1,1), "Partido A vs B", "B1", 1.80, 100, "Ganado"],
    [date(2026,1,2), "Partido C vs D", "C2", 2.40, 50, "Perdido"],
    [date(2026,1,3), "Partido E vs F", "E1", 1.65, 75, "Ganado"],
    [date(2026,1,4), "Torneo G", "G-DC", 3.10, 30, "Perdido"],
    [date(2026,1,5), "Exhibición H", "H1", 2.00, 120, "Nulo"],
]
for row in sample:
    ws.append(row)

maxrow = ws.max_row
for r in range(2, maxrow + 1):
    ws[f'G{r}'] = f'=IF(F{r}="Ganado",E{r}*D{r},IF(F{r}="Nulo",E{r},0))'
    ws[f'H{r}'] = f'=G{r}-E{r}'
    ws[f'I{r}'] = f'=IF(E{r}=0,0,H{r}/E{r})'

header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True, color='000000')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

widths = [12, 30, 18, 10, 14, 12, 12, 14, 10]
for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], widths):
    ws.column_dimensions[col].width = w

thin = Side(border_style='thin', color='000000')
for row in ws.iter_rows(min_row=1, max_row=maxrow, min_col=1, max_col=9):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

green_dxf = DifferentialStyle(fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
red_dxf = DifferentialStyle(fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))

ws.conditional_formatting.add(f'H2:H{maxrow}', CellIsRule(operator='greaterThan', formula=['0'], dxf=green_dxf))
ws.conditional_formatting.add(f'H2:H{maxrow}', CellIsRule(operator='lessThan', formula=['0'], dxf=red_dxf))

ws.conditional_formatting.add(f'A2:I{maxrow}', FormulaRule(formula=['$F2="Ganado"'], dxf=green_dxf))
ws.conditional_formatting.add(f'A2:I{maxrow}', FormulaRule(formula=['$F2="Perdido"'], dxf=red_dxf))

dv = DataValidation(type='list', formula1='"Ganado,Perdido,Nulo"', allow_blank=False)
ws.add_data_validation(dv)
dv.ranges.append(f'F2:F{maxrow + 100}')

ws['K1'] = 'Métrica'
ws['L1'] = 'Valor'
metrics = [
    ('Total Apuestas', f'=COUNTA(A2:A{maxrow + 100})'),
    ('Apuestas Ganadas', f'=COUNTIF(F2:F{maxrow + 100},"Ganado")'),
    ('Apuestas Perdidas', f'=COUNTIF(F2:F{maxrow + 100},"Perdido")'),
    ('Apuestas Nulas', f'=COUNTIF(F2:F{maxrow + 100},"Nulo")'),
    ('% Acierto', f'=IF(L2=0,0,L3/L2)'),
    ('ROI Promedio', f'=IF(L2=0,0,SUM(H2:H{maxrow + 100})/SUM(E2:E{maxrow + 100}))'),
    ('Ganancia Total', f'=SUM(H2:H{maxrow + 100})'),
]
for idx, (name, formula) in enumerate(metrics, start=2):
    ws[f'K{idx}'] = name
    ws[f'L{idx}'] = formula
    ws[f'K{idx}'].font = Font(bold=True)

ndash = wb.create_sheet('Dashboard')
dash['A1'] = 'Dashboard de Rendimiento'
dash['A1'].font = Font(size=16, bold=True)

ndash['A3'] = 'Etiqueta'; ndash['B3'] = 'Valores'
ndash['A4'] = 'Ganancia Total'; ndash['B4'] = '=Registro Apuestas!L8'
ndash['A5'] = 'ROI Promedio'; ndash['B5'] = '=Registro Apuestas!L6'
ndash['A6'] = '% Acierto'; ndash['B6'] = '=Registro Apuestas!L5'

chart = BarChart()
chart.title = 'Rendimiento General'
chart.x_axis.title = 'Métricas'
chart.y_axis.title = 'Valor'
cat = Reference(dash, min_col=1, min_row=4, max_row=6)
val = Reference(dash, min_col=2, min_row=4, max_row=6)
chart.add_data(val, titles_from_data=False)
chart.set_categories(cat)
chart.shape = 4
dash.add_chart(chart, 'D3')

wb.save('ControlLudMejorado.xlsx')
print('Creado ControlLudMejorado.xlsx con mejoras.')
